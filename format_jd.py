# dashboard_formatter.py
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows

class DashboardFormatter:
    def __init__(self, file_path, title=None, publication_date=None):
        self.workbook = load_workbook(file_path)
        self.file_path = file_path
        self.title = title
        self.publication_date = publication_date

    def formatter_jd(self, df):
            sheet = self.workbook["Hiring Details"]
            sheet["C3"] = self.title
            sheet["C3"].font = Font(name="Calibri", size=10, italic=True)
            # Paste DataFrame into the sheet starting from row 7
            font_style = Font(name="Calibri", size=10)  # Change to your preferred font and size

            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=6):
                for col_idx, value in enumerate(row_data, start=3):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = font_style
            # Function to center align specific columns from row 7 onwards
            def center_align_column(sheet, column_letter, start_row=6):
                for cell in sheet[column_letter]:
                    if cell.row >= start_row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            def left_align_column(sheet, column_letter, start_row=6):
                for cell in sheet[column_letter]:
                    if cell.row >= start_row:
                        cell.alignment = Alignment(horizontal="left")
            # Center align columns C, J, L, M, N from row 7 onwards
            for col in [get_column_letter(i) for i in range(column_index_from_string('C'), column_index_from_string('O') + 1)]:
                center_align_column(sheet, col)

            for col in ['E', 'H' ,'I','J','K','L','M','N']:
                left_align_column(sheet, col)

            # Function to adjust column width based on the longest content
            def adjust_column_width(sheet, column_letter, padding=0):
                max_length = 0
                for cell in sheet[column_letter]:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column_letter].width = max_length + padding

            adjust_column_width(sheet, 'F')  # adjust padding if needed

            # Center align cells with a "-" value
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == "-":
                        cell.alignment = Alignment(horizontal="center")
            # Border Processing
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

            start_column = 'C'
            start_row = 6

            num_columns = len(df.columns) - 1

            end_column = get_column_letter(column_index_from_string(start_column) + num_columns)
            for row_index in range(start_row, start_row + len(df)):
                for column_index in range(column_index_from_string(start_column), column_index_from_string(end_column) + 1):
                    cell = sheet.cell(row=row_index, column=column_index)
                    cell.border = border

            for cell in sheet['C']:
                if cell.row < 6 or 6 + len(df) < cell.row:
                    continue
                cell.border = Border(left=Side(style='medium'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            for cell in sheet['O']:
                if cell.row < 6 or 6 + len(df) < cell.row:
                    continue
                cell.border = Border(left=Side(style='thin'),right=Side(style='medium'),top=Side(style='thin'),bottom=Side(style='thin'))

            for column_index in range(column_index_from_string(start_column), column_index_from_string(end_column) + 1):
                if column_index == 3:
                    cell = sheet.cell(row=5 + len(df), column=column_index)
                    cell.border = Border(left=Side(style='medium'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='medium'))
                elif column_index == 15:
                    cell = sheet.cell(row=5 + len(df), column=column_index)
                    cell.border = Border(left=Side(style='thin'),right=Side(style='medium'),top=Side(style='thin'),bottom=Side(style='medium'))
                else:
                    cell = sheet.cell(row=5 + len(df), column=column_index)
                    cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='medium'))

    def save(self, output_path):
        self.workbook.save(output_path)
